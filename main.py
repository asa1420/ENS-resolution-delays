import requests
import time
import openpyxl
import datetime
now = datetime.datetime.now()
wb = openpyxl.load_workbook("C:/Users/Abdullah/go/src/ENS/measurement.xlsx")
sh1 = wb["Sheet1"]
ENSdomain = ["http://ethereum.eth.link","http://almonit.eth.link","http://pepesza.eth.link","http://alex.eth.link","http://bitcoingenerator.eth.link"]
for i in range(5):
    start = time.time()
    req = requests.get(ENSdomain[i])
    end = time.time()
    delay = end - start
    print(delay)
    sh1.cell(9,i+16).value = delay
sh1.cell(9,15).value = now.strftime("%H:%M:%S")
wb.save("C:/Users/Abdullah/go/src/ENS/measurement.xlsx")