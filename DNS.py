import requests
import time
import openpyxl
import datetime
now = datetime.datetime.now()
wb = openpyxl.load_workbook("C:/Users/Abdullah/go/src/ENS/measurement.xlsx")
sh1 = wb["Sheet1"]
domains = ["https://docs.ipfs.io/how-to/websites-on-ipfs/link-a-domain/#ethereum-naming-service-ens","https://timetable.ucl.ac.uk/tt/persTimet.do","https://www.youtube.com/","https://moodle.ucl.ac.uk/login/index.php","https://drive.google.com/drive/my-drive"]
for i in range(5):
    start = time.time()
    req = requests.get(domains[i])
    end = time.time()
    delay = end - start
    print(delay)
    sh1.cell(13,i+23).value = delay
sh1.cell(13,22).value = now.strftime("%H:%M:%S")
wb.save("C:/Users/Abdullah/go/src/ENS/measurement.xlsx")